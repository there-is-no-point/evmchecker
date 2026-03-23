import asyncio
import csv
import json
import os
import random
import time
from datetime import datetime
from decimal import Decimal

import openpyxl

import aiohttp
import questionary
from rich import box
from rich.console import Console
from rich.progress import BarColumn, Progress, SpinnerColumn, TaskID
from rich.table import Table
from web3 import AsyncHTTPProvider, AsyncWeb3

from networks import ERC20_ABI, MULTICALL_ABI, config as _raw_config
import config as cfg

# ─────────────────────────────────────────────────────────────────────────────
# Network filtering by NETWORK_MODE
# ─────────────────────────────────────────────────────────────────────────────

def _build_config(mode: str) -> dict:
    """
    Возвращает отфильтрованный словарь сетей согласно NETWORK_MODE:
      "all"      — все сети
      "mainnet"  — только основные (без флага testnet: True)
      "testnet"  — только тестовые (с флагом testnet: True)
    """
    if mode == "all":
        return dict(_raw_config)
    elif mode == "testnet":
        return {k: v for k, v in _raw_config.items() if v.get("testnet", False)}
    else:  # "mainnet" — default
        return {k: v for k, v in _raw_config.items() if not v.get("testnet", False)}

config = _build_config(cfg.NETWORK_MODE)

# ─────────────────────────────────────────────────────────────────────────────
# Constants (pulled from config.py)
# ─────────────────────────────────────────────────────────────────────────────
console = Console()
MAX_CONCURRENT        = cfg.MAX_CONCURRENT
MIN_CONCURRENT        = cfg.MIN_CONCURRENT
MIN_VALUE_TO_DISPLAY  = cfg.MIN_VALUE_TO_DISPLAY
RETRY_PER_RPC         = cfg.RETRY_PER_RPC
COMPARE_PRICES        = cfg.COMPARE_PRICES
PRICE_DIVERGENCE_THRESHOLD = Decimal(cfg.PRICE_DIVERGENCE_THRESHOLD)

# Global state
SELECTED_CHAIN = None
NATIVE_TOKEN   = None

ALL_NETWORKS_CHOICE = "[ All Networks ]"


def read_file(path: str) -> list[str]:
    with open(path) as f:
        return [line.strip() for line in f if line.strip()]


def load_addresses() -> list[str]:
    addresses = read_file("addresses.txt")
    if not addresses:
        console.print("[red]Ошибка:[/red] файл [yellow]addresses.txt[/yellow] пуст — добавь адреса кошельков (по одному на строку).")
        exit(1)
    return addresses


def select_chain() -> bool:
    """
    Показывает меню выбора.
    Возвращает True  — одна сеть выбрана, SELECTED_CHAIN / NATIVE_TOKEN установлены.
    Возвращает False — выбран режим All Networks.
    """
    global SELECTED_CHAIN, NATIVE_TOKEN

    if not config:
        mode_label = {"all": "All", "mainnet": "Mainnet", "testnet": "Testnet"}.get(cfg.NETWORK_MODE, cfg.NETWORK_MODE)
        console.print(f"[red]No networks available for NETWORK_MODE = \"{cfg.NETWORK_MODE}\".[/red]")
        console.print(f"[yellow]Check networks.py — make sure at least one chain matches the \"{mode_label}\" filter.[/yellow]")
        exit(1)

    choices = [ALL_NETWORKS_CHOICE] + list(config.keys())

    chain = questionary.autocomplete(
        "Select chain (type to search):", choices=choices, match_middle=False
    ).ask()

    if chain is None:
        exit()

    if chain == ALL_NETWORKS_CHOICE:
        return False

    try:
        SELECTED_CHAIN = chain.lower()
        NATIVE_TOKEN = config[SELECTED_CHAIN]["symbol"]
        return True
    except KeyError:
        console.print(f"[yellow]Chain {chain} not found[/yellow]")
        exit()


def set_chain(chain: str) -> None:
    """Устанавливает активную сеть глобально (используется в All Networks)."""
    global SELECTED_CHAIN, NATIVE_TOKEN
    SELECTED_CHAIN = chain
    NATIVE_TOKEN = config[chain]["symbol"]


def get_web3(proxies: list[str], rpc_url: str | None = None) -> AsyncWeb3:
    assert SELECTED_CHAIN is not None

    if rpc_url is None:
        rpc_url = random.choice(config[SELECTED_CHAIN]["rpc"])
    kwargs = {"proxy": f"http://{random.choice(proxies)}"} if proxies else {}
    return AsyncWeb3(AsyncHTTPProvider(rpc_url, request_kwargs=kwargs))


# ─────────────────────────────────────────────────────────────────────────────
# Price sources  (берутся целиком из config.py)
# ─────────────────────────────────────────────────────────────────────────────
PRICE_SOURCES = cfg.PRICE_SOURCES


async def _fetch_price_from_source(session: aiohttp.ClientSession, source: dict) -> Decimal | None:
    """Запрашивает цену с одной биржи. Возвращает Decimal или None при ошибке."""
    url = source["url"].format(symbol=NATIVE_TOKEN)
    try:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=5)) as response:
            if response.status == 200:
                data = await response.json()
                price = source["parse"](data)
                if price > 0:
                    return price
    except Exception:
        pass
    return None


async def get_token_price() -> Decimal:
    """
    Получает USD-цену нативного токена.

    COMPARE_PRICES = False (по умолчанию):
        Опрашивает биржи по порядку, возвращает первую валидную цену.

    COMPARE_PRICES = True:
        Опрашивает все биржи параллельно, считает среднюю цену.
        Если разброс между min и max превышает PRICE_DIVERGENCE_THRESHOLD —
        выводит предупреждение и показывает цену каждой биржи.
    """
    async with aiohttp.ClientSession() as session:

        if not COMPARE_PRICES:
            for source in PRICE_SOURCES:
                price = await _fetch_price_from_source(session, source)
                if price is not None:
                    return price
                console.print(
                    f"[yellow]{source['name']}: no price for {NATIVE_TOKEN}, trying next...[/yellow]"
                )
            console.print(f"[red]Could not fetch price for {NATIVE_TOKEN} from any exchange[/red]")
            return Decimal(0)

        else:
            tasks = [_fetch_price_from_source(session, source) for source in PRICE_SOURCES]
            raw_results = await asyncio.gather(*tasks)

            prices: list[tuple[str, Decimal]] = [
                (PRICE_SOURCES[i]["name"], price)
                for i, price in enumerate(raw_results)
                if price is not None
            ]

            if not prices:
                console.print(f"[red]Could not fetch price for {NATIVE_TOKEN} from any exchange[/red]")
                return Decimal(0)

            values = [p for _, p in prices]
            avg_price = sum(values) / len(values)

            if len(values) > 1:
                min_price = min(values)
                max_price = max(values)
                divergence = (max_price - min_price) / min_price

                if divergence > PRICE_DIVERGENCE_THRESHOLD:
                    console.print(
                        f"[bold yellow]⚠ Price divergence for {NATIVE_TOKEN}: "
                        f"{float(divergence * 100):.2f}% (threshold: {float(PRICE_DIVERGENCE_THRESHOLD * 100):.0f}%)[/bold yellow]"
                    )
                    for name, price in prices:
                        console.print(f"  [yellow]{name:<10}[/yellow] ${float(price):.6f}")
                    console.print(f"  [bold]Average  [/bold] ${float(avg_price):.6f}")

            return avg_price


# ─────────────────────────────────────────────────────────────────────────────
# Balance fetching
# ─────────────────────────────────────────────────────────────────────────────

async def get_balance_from_multicall(address: str, proxies: list[str]) -> dict:
    assert SELECTED_CHAIN is not None

    checksum_address = AsyncWeb3.to_checksum_address(address)
    multicall_address = AsyncWeb3.to_checksum_address(config[SELECTED_CHAIN]["multicall"])
    token_list = list(config[SELECTED_CHAIN]["tokens"].items())
    default_token_balances = {token_name: 0 for token_name, _ in token_list}

    rpc_pool = list(config[SELECTED_CHAIN]["rpc"])
    first_rpc = random.choice(rpc_pool)
    rpc_list = [first_rpc] + [r for r in rpc_pool if r != first_rpc]

    for rpc_url in rpc_list:
        for attempt in range(RETRY_PER_RPC):
            try:
                web3 = get_web3(proxies, rpc_url=rpc_url)
                multicall = web3.eth.contract(multicall_address, abi=MULTICALL_ABI)
                calls = []

                eth_balance_call = multicall.encodeABI("getEthBalance", [checksum_address])
                calls.append({"target": multicall_address, "allowFailure": False, "callData": eth_balance_call})

                for token_name, token_address in token_list:
                    token_address = web3.to_checksum_address(token_address)
                    token_contract = web3.eth.contract(token_address, abi=ERC20_ABI)

                    balance_call = token_contract.encodeABI("balanceOf", [checksum_address])
                    calls.append({"target": token_address, "allowFailure": False, "callData": balance_call})

                    decimals_call = token_contract.encodeABI("decimals", [])
                    calls.append({"target": token_address, "allowFailure": False, "callData": decimals_call})

                results = await multicall.functions.aggregate3(calls).call()

                balance_wei = web3.to_int(results[0][1])
                balance_eth = web3.from_wei(balance_wei, "ether")

                token_balances = {}
                for i, (token_name, _) in enumerate(token_list):
                    balance_result = results[1 + i * 2]
                    decimals_result = results[1 + i * 2 + 1]

                    balance = web3.to_int(balance_result[1])
                    decimals = web3.to_int(decimals_result[1])

                    token_balances[token_name] = float(Decimal(balance) / Decimal(10**decimals))

                await asyncio.sleep(random.uniform(0.05, 0.15))
                nonce = await web3.eth.get_transaction_count(checksum_address)

                return {"address": address, NATIVE_TOKEN: float(balance_eth), **token_balances, "tx_count": int(nonce)}

            except Exception as e:
                is_rate_limit = (
                    "429" in str(e)
                    or "Too Many Requests" in str(e).lower()
                    or "call rate limit exhausted" in str(e).lower()
                )
                if is_rate_limit:
                    await asyncio.sleep(0.5 * (2**attempt))
                if attempt == RETRY_PER_RPC - 1:
                    short_rpc = rpc_url.split("//")[-1][:40]
                    console.print(
                        f"[yellow]RPC failed ({short_rpc}), switching to next...[/yellow]"
                    )

    console.print(f"[red]All RPCs failed for[/red] {address}")
    return {"address": address, NATIVE_TOKEN: 0.0, **default_token_balances, "tx_count": int(0)}


async def check_balances(addresses: list[str], proxies: list[str], progress: Progress, task_id: TaskID) -> list[dict]:
    max_concurrent = MIN_CONCURRENT if not proxies else MAX_CONCURRENT

    semaphore = asyncio.Semaphore(max_concurrent)
    completed = 0

    async def throttled_get_balance(address: str) -> dict:
        nonlocal completed
        async with semaphore:
            result = await get_balance_from_multicall(address, proxies)
            completed += 1
            progress.update(task_id, completed=completed)
            return result

    tasks = [throttled_get_balance(address) for address in addresses]
    results = await asyncio.gather(*tasks)
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Output
# ─────────────────────────────────────────────────────────────────────────────

def format_value(value: str | float | int) -> str | int:
    if value == 0:
        return ""
    if isinstance(value, float) and 0 < value < MIN_VALUE_TO_DISPLAY:
        return "~ 0"
    if isinstance(value, float):
        return f"{value:.8f}"
    return value


def print_table(results: list[dict], token_price: Decimal, native_token: str | None = None) -> None:
    # Если native_token не передан явно — берём из глобального (режим одной сети)
    nt = native_token if native_token is not None else NATIVE_TOKEN

    table = Table(box=box.ASCII_DOUBLE_HEAD, show_footer=True)
    headers = list(results[0].keys())

    totals = {}
    for header in headers:
        if header not in ["address"]:
            totals[header] = sum(entry[header] for entry in results if isinstance(entry[header], (int, float)))

    for header in headers:
        footer_text = ""
        if header == "address":
            footer_text = "Total"
        elif header == "USD" and token_price:
            total_usd = Decimal(str(totals.get(nt, 0))) * token_price
            footer_text = f"{total_usd:.2f}" if total_usd > 0 else ""
        elif header == "tx_count":
            footer_text = f"{int(totals[header])}" if totals[header] > 0 else ""
        elif header in totals:
            footer_text = f"{totals[header]:.6f}" if totals[header] > 0 else ""

        table.add_column(
            header, footer=footer_text, justify="right" if header != "address" else "left", footer_style="bold green"
        )

    for result in results:
        formatted_values = [f"[cyan]{format_value(value)}[/cyan]" for value in result.values()]
        table.add_row(*formatted_values)

    console.print(table)


def write_to_csv(results: list[dict], filename: str) -> None:
    headers = list(results[0].keys())
    formatted_values = [{key: str(format_value(value)) for key, value in entry.items()} for entry in results]

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(formatted_values)


def write_to_xlsx(results: list[dict], filename: str) -> None:
    headers = list(results[0].keys())

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row
    ws.append(headers)
    header_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="1E1E2E")
    header_font = openpyxl.styles.Font(bold=True, color="00FF99")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for entry in results:
        ws.append([str(format_value(v)) for v in entry.values()])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(filename)


def write_to_json(results: list[dict], filename: str) -> None:
    formatted = [{key: str(format_value(value)) for key, value in entry.items()} for entry in results]
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(formatted, f, ensure_ascii=False, indent=2)


def select_output_format() -> str:
    """Спрашивает пользователя в каком формате сохранить результаты."""
    fmt = questionary.select(
        "Select output format:",
        choices=["CSV", "Excel (.xlsx)", "JSON"],
    ).ask()
    if fmt is None:
        exit()
    return fmt


def save_results(results: list[dict], chain: str, fmt: str) -> str:
    """Сохраняет результаты в выбранном формате, возвращает путь к файлу."""
    os.makedirs("results", exist_ok=True)
    ts = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    ext_map = {"CSV": "csv", "Excel (.xlsx)": "xlsx", "JSON": "json"}
    ext = ext_map[fmt]
    filename = f"results/{chain}-{ts}.{ext}"

    if fmt == "CSV":
        write_to_csv(results, filename)
    elif fmt == "Excel (.xlsx)":
        write_to_xlsx(results, filename)
    elif fmt == "JSON":
        write_to_json(results, filename)

    return filename


# ─────────────────────────────────────────────────────────────────────────────
# Run modes
# ─────────────────────────────────────────────────────────────────────────────

async def run_single_chain(
    chain: str,
    addresses: list[str],
    proxies: list[str],
    fmt: str,
    progress: Progress | None = None,
    chains_task: TaskID | None = None,
) -> tuple[list[dict], Decimal, float]:
    """
    Проверяет адреса для одной сети.
    Возвращает (results, token_price, elapsed).
    Сохранение и вывод таблицы — на усмотрение вызывающего кода.
    """
    set_chain(chain)

    inner_progress = Progress(
        SpinnerColumn(),
        "[progress.description]{task.description}",
        BarColumn(),
        "[progress.percentage]{task.percentage:>3.0f}%",
        transient=True,
        console=console,
    )
    with inner_progress:
        task = inner_progress.add_task(
            f"Checking [cyan]{len(addresses)}[/cyan] addresses on {chain.title()} ",
            total=len(addresses),
        )

        start_time = time.time()
        results = await check_balances(addresses, proxies, progress=inner_progress, task_id=task)
        token_price = await get_token_price()

        if token_price:
            for result in results:
                result["USD"] = float(Decimal(str(result[NATIVE_TOKEN])) * token_price)

    elapsed = time.time() - start_time

    # Обновляем общий прогресс-бар (если передан)
    if progress is not None and chains_task is not None:
        progress.advance(chains_task)

    return results, token_price, elapsed, NATIVE_TOKEN


async def main(fmt: str) -> None:
    """Режим одной сети."""
    assert SELECTED_CHAIN is not None
    addresses = load_addresses()
    proxies = read_file("proxies.txt")
    results, token_price, elapsed, native_token = await run_single_chain(SELECTED_CHAIN, addresses, proxies, fmt)
    print_table(results, token_price, native_token)
    filename = save_results(results, SELECTED_CHAIN, fmt)
    console.print(f"Execution time: {elapsed:.2f} seconds")
    console.print(f"Results saved to [green]{filename}[/green]\n")


def _is_empty(results: list[dict]) -> bool:
    """Возвращает True если все балансы во всех записях нулевые."""
    for entry in results:
        for key, val in entry.items():
            if key == "address":
                continue
            if isinstance(val, (int, float)) and val > 0:
                return False
    return True


def _print_summary_table(all_results: dict[str, tuple[list[dict], Decimal]]) -> None:
    """Выводит сводную таблицу: адрес / сеть / USD-баланс нативного токена."""
    table = Table(
        title="[bold cyan]All Networks — Summary[/bold cyan]",
        box=box.ASCII_DOUBLE_HEAD,
        show_footer=True,
    )
    table.add_column("Address", justify="left")
    table.add_column("Network", justify="left")
    table.add_column("Native", justify="right")
    table.add_column("USD", justify="right", footer_style="bold green")

    total_usd = Decimal(0)
    for chain, (results, token_price, native_token) in all_results.items():
        native_symbol = config[chain]["symbol"]
        for entry in results:
            native_bal = entry.get(native_symbol, 0.0)
            usd_val = entry.get("USD", 0.0)
            total_usd += Decimal(str(usd_val))
            table.add_row(
                f"[cyan]{entry['address']}[/cyan]",
                chain.title(),
                f"{native_bal:.8f}" if native_bal else "",
                f"{usd_val:.2f}" if usd_val else "",
            )

    table.columns[3].footer = f"{total_usd:.2f}"
    console.print(table)


def _save_all_networks_excel(all_results: dict[str, tuple[list[dict], Decimal]], ts: str) -> str:
    """Сохраняет все сети в один Excel-файл, каждая сеть — отдельный лист."""
    os.makedirs("results", exist_ok=True)
    filename = f"results/all_networks-{ts}.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # убираем дефолтный пустой лист

    header_fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="1E1E2E")
    header_font = openpyxl.styles.Font(bold=True, color="00FF99")

    for chain, (results, *_) in all_results.items():
        ws = wb.create_sheet(title=chain.title()[:31])  # Excel ограничивает имя листа 31 символом
        headers = list(results[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for entry in results:
            ws.append([str(format_value(v)) for v in entry.values()])
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(filename)
    return filename


def _save_all_networks_json(all_results: dict[str, tuple[list[dict], Decimal]], ts: str) -> str:
    """Сохраняет все сети в один JSON-файл."""
    os.makedirs("results", exist_ok=True)
    filename = f"results/all_networks-{ts}.json"
    output = {}
    for chain, (results, *_) in all_results.items():
        output[chain] = [{key: str(format_value(val)) for key, val in entry.items()} for entry in results]
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    return filename


def _save_all_networks_csv(all_results: dict[str, tuple[list[dict], Decimal]], ts: str) -> list[str]:
    """Сохраняет каждую сеть в отдельный CSV (стандартное поведение)."""
    files = []
    for chain, (results, *_) in all_results.items():
        filename = f"results/{chain}-{ts}.csv"
        write_to_csv(results, filename)
        files.append(filename)
    return files


async def main_all_networks(fmt: str) -> None:
    """
    Режим All Networks:
    - Последовательный обход сетей (глобальный стейт NATIVE_TOKEN/SELECTED_CHAIN не конфликтует)
    - Общий прогресс-бар по сетям
    - Пропуск пустых сетей
    - Сводная таблица в конце
    - Один файл для Excel/JSON, отдельные CSV для каждой сети
    """
    addresses = load_addresses()
    proxies = read_file("proxies.txt")
    chains = list(config.keys())
    total_start = time.time()
    ts = datetime.now().strftime("%d_%m_%Y_%H_%M_%S")

    mode_label = {"all": "All", "mainnet": "Mainnet only", "testnet": "Testnet only"}.get(
        cfg.NETWORK_MODE, cfg.NETWORK_MODE
    )
    console.print(
        f"\n[bold cyan]All Networks mode[/bold cyan] "
        f"([dim]{mode_label}[/dim]) — "
        f"[yellow]{len(chains)}[/yellow] chains, "
        f"[yellow]{len(addresses)}[/yellow] addresses\n"
    )

    all_results: dict[str, tuple[list[dict], Decimal]] = {}
    skipped: list[str] = []

    # Общий прогресс-бар по сетям (остаётся виден пока идёт обход)
    overall_progress = Progress(
        SpinnerColumn(),
        "[progress.description]{task.description}",
        BarColumn(),
        "[progress.percentage]{task.percentage:>3.0f}%",
        "[yellow]{task.completed}/{task.total} chains[/yellow]",
        console=console,
        transient=False,
    )

    with overall_progress:
        chains_task = overall_progress.add_task(
            "[cyan]All Networks[/cyan]", total=len(chains)
        )

        for chain in chains:
            try:
                results, token_price, elapsed, native_token = await run_single_chain(
                    chain, addresses, proxies, fmt,
                    progress=overall_progress, chains_task=chains_task,
                )
                if _is_empty(results):
                    skipped.append(chain)
                    overall_progress.console.print(f"[dim]  {chain.title()} — skipped (empty)[/dim]")
                else:
                    all_results[chain] = (results, token_price, native_token)
                    overall_progress.console.print(
                        f"  [bold]{chain.title()}[/bold] — {len(results)} addresses — {elapsed:.2f}s"
                    )
            except Exception as e:
                overall_progress.advance(chains_task)
                overall_progress.console.print(f"[red]  Error on {chain}:[/red] {e}")

    if skipped:
        console.print(f"\n[dim]Skipped empty networks: {', '.join(skipped)}[/dim]")

    if not all_results:
        console.print("\n[yellow]No balances found across all networks.[/yellow]")
        return

    # Вывод таблиц по каждой непустой сети
    for chain, (results, token_price, native_token) in all_results.items():
        console.print(f"\n[bold yellow]{chain.title()}[/bold yellow]")
        print_table(results, token_price, native_token)

    # Сводная таблица
    _print_summary_table(all_results)

    # Сохранение
    os.makedirs("results", exist_ok=True)
    if fmt == "Excel (.xlsx)":
        filename = _save_all_networks_excel(all_results, ts)
        console.print(f"\nAll results saved to [green]{filename}[/green] (one sheet per network)")
    elif fmt == "JSON":
        filename = _save_all_networks_json(all_results, ts)
        console.print(f"\nAll results saved to [green]{filename}[/green]")
    else:  # CSV — отдельный файл на каждую сеть
        files = _save_all_networks_csv(all_results, ts)
        console.print(f"\nAll results saved to [green]results/[/green] ({len(files)} files)")

    elapsed_total = time.time() - total_start
    console.print(f"[bold green]All done![/bold green] Total time: {elapsed_total:.2f} seconds\n")


if __name__ == "__main__":
    try:
        single_mode = select_chain()
        fmt = select_output_format()
        if single_mode:
            asyncio.run(main(fmt))
        else:
            asyncio.run(main_all_networks(fmt))
    except KeyboardInterrupt:
        pass
